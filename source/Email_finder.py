# %% Email Finder{{{

# Imports
from time import sleep
from bs4 import BeautifulSoup
from openpyxl.cell import WriteOnlyCell  # To edit cell attributes
from openpyxl.styles import PatternFill  # Cell Color Fills

"""
EmailFinder.py
=========================
Creates the cell types used if the email is requested.

Defines the function EmailFinder.

"""

# Cells/Colors that will be used a lot:

''''COLORS for highlighting'''
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

orangeFill = PatternFill(start_color='FFC000',
                         end_color='FFC000',
                         fill_type='solid')

yellowFill = PatternFill(start_color='FDF96F',
                         end_color='FDF96F',
                         fill_type='solid')

"EMAIL NOT FINDABLE CELl"
No_email_cell = WriteOnlyCell(ws, value="Email Not Found/Findable")
No_email_cell.fill = orangeFill

"EMAIL NOT SEARCHED FOR CELL"
No_email_cell2 = WriteOnlyCell(ws, value="Email Not Searched For")

"NAME NOT PARSABLE CELl"
No_name_cell = WriteOnlyCell(ws, value="No Instructor")
No_name_cell.fill = redFill

"COURSE NOT OFFERED CELL"
No_course_cell = WriteOnlyCell(ws, value="No Sections Offered")
No_course_cell.fill = yellowFill


def EmailFinder(NAME_TUPLE, DRIVER, EmailVar):
    """
    Takes the name as tuple of first and last name and searches using the
    selenium driver on the directory Uses the predefined cell types to give
    some highlighting where needed.

    :params tuple NAME_TUPLE: A tuple of the first and last name.
    :params selenium.driver DRIVER: The selenium browser driver used.
    :returns: A list of the name as a string, then the relevant cell style.
    :rtype: list
    """

    firstName = NAME_TUPLE[0]
    lastName = NAME_TUPLE[1]
    if firstName == 'No Instructor':
        return [No_name_cell, No_email_cell2]
    else:
        if EmailVar is True:
            DRIVER.get('https://dir.unc.edu/search/'+firstName+'%20'+lastName)
            sleep(1)
            if " If the error indicates a time limit" in DRIVER.page_source or "This page isn’t working" in DRIVER.page_source:
                sleep(5)
                return EF(NAME_TUPLE, DRIVER, No_email_cell,
                          No_email_cell2, No_course_cell,
                          No_name_cell, EmailVar)
            else:
                # Check to make sure something was found
                if "no results" not in DRIVER.page_source:
                    # save the page
                    soup_file = DRIVER.page_source
                    soup = BeautifulSoup(soup_file, 'html.parser')

                    email_link = soup.select('a[href^=mailto]')

                    try:
                        href = email_link[0]['href']

                        try:
                            str1, email = href.split(':')
                        except ValueError:
                            return [firstName + " " + lastName, No_email_cell]

                        # if it does work return the email
                        return [firstName + " " + lastName, email]
                    except IndexError:
                        return [firstName + " " + lastName, No_email_cell]
                else:
                    return [firstName + " " + lastName, No_email_cell]
        else:
            return [firstName + " " + lastName, No_email_cell2]


# }}}
