# %% Excel Start Function
def ESF():
    import openpyxl  # Creates the excel sheets and workbooks that I will use
    from openpyxl.cell import WriteOnlyCell  # To edit cell attributes
    from openpyxl.styles import PatternFill  # Cell Color Fills
    # Start a WRITE ONLY workbook to keep ram usage low
    wb = openpyxl.Workbook(write_only=True)

    # need to create a sheet to start as the workbook is write_only
    ws = wb.create_sheet()

    # Cells/Colors that will be used a lot:

    ''''COLORS for highlighting'''
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    orangeFill = PatternFill(start_color='FFC000',
                             end_color='FFC000',
                             fill_type='solid')

    yellowFill = PatternFill(start_color='FDF96F',
                             end_color='FDF96F',
                             fill_type='solid')

    "EMAIL NOT FINDABLE CELl"
    No_email_cell = WriteOnlyCell(ws, value="Email Not Found/Findable")
    No_email_cell.fill = orangeFill
    "EMAIL NOT SEARCHED FOR CELL"
    No_email_cell2 = WriteOnlyCell(ws, value="Email Not Searched For")

    "NAME NOT PARSABLE CELl"
    No_name_cell = WriteOnlyCell(ws, value="No Instructor")
    No_name_cell.fill = redFill

    "COURSE NOT OFFERED CELL"
    No_course_cell = WriteOnlyCell(ws, value="No Sections Offered")
    No_course_cell.fill = yellowFill

    return ([wb, ws], [No_email_cell, No_email_cell2, No_course_cell, No_name_cell])
# }}}
